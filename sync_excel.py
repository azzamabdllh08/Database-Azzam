from pathlib import Path
import json
import openpyxl

source = Path("source.xlsx")
output = Path("data.js")

if not source.exists():
    raise SystemExit("source.xlsx tidak ditemukan.")

wb = openpyxl.load_workbook(source, data_only=True, read_only=True)

# File Bos menggunakan sheet Data_Master.
sheet_name = "Data_Master"
if sheet_name not in wb.sheetnames:
    raise SystemExit(f"Sheet {sheet_name!r} tidak ditemukan. Sheet tersedia: {wb.sheetnames}")

ws = wb[sheet_name]
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
headers = [str(x).strip() if x is not None else "" for x in header_row]

needed = ["Site", "Category", "Description", "Tag_No", "Status", "Survey_Date", "Surveyed"]
missing = [x for x in needed if x not in headers]
if missing:
    raise SystemExit(f"Kolom tidak ditemukan: {missing}")

idx = {h:i for i,h in enumerate(headers)}
records = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(v is not None for v in row):
        continue
    item = {}
    for h in needed:
        v = row[idx[h]]
        if hasattr(v, "isoformat"):
            v = v.isoformat()
        item[h] = v
    records.append(item)

output.write_text(
    "window.SITE_DATA=" + json.dumps(records, ensure_ascii=False, separators=(",", ":")) + ";",
    encoding="utf-8"
)

print(f"Generated {output} with {len(records)} records.")
