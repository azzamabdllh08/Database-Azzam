import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

SOURCE = Path("Input RBI.xlsx")
OUTPUT = Path("data.js")

# Row 3 contains the database headers; rows 5 onward contain assets.
HEADER_ROW = 3
DATA_START_ROW = 5

FIELDS = {
    "No": "no", "Ref. No.": "refNo", "P/F": "pf", "Comodity Code": "commodity",
    "System Name": "system", "Line No.": "lineNo", "Item Description": "description",
    "Originating": "originating", "Terminating": "terminating", "PID": "pid", "Status": "status",
    "Commision Date": "commissionDate", "Service Fluid": "serviceFluid", "Primary Fluid": "primaryFluid",
    "Most Volatile Fluid": "volatileFluid", "Toxic\nConstituent": "toxic", "Flow Rate\n(ft3/h)": "flowRate",
    "DP \n(psig)": "designPressure", "DT \n(0F)": "designTemp", "Max. OP\n(psig)": "maxOpPressure",
    "Max. OT \n(0F)": "maxOpTemp", "Material": "material", "NPS": "nps", "OD \n(inch)": "od",
    "Class": "class", "Rating": "rating", "SCH.": "schedule", "Nom. THK.\n(mm)": "nomThk",
    "CA \n(mm)": "ca", "Design Code": "designCode", "Insulation": "insulation",
    "Isometric As Built": "isometric", "Assessment Date": "assessmentDate", "Last Insp. Date": "lastInspectionDate",
    "Inspection Method": "inspectionMethod", "ROI Number": "roi", "Visual Finding": "visualFinding",
    "Thickness Measurement": "thicknessMeasurement", "Internal issue\n(Yes/No)": "internalIssue",
    "External issue\n(Yes/No)": "externalIssue", "Repair\n(Yes/No)": "repair",
    "Repair Description": "repairDescription", "LF1": "lf1", "LF3": "lf3", "LF4": "lf4", "LF5": "lf5",
    "LF6": "lf6", "LF7": "lf7", "CF5": "cf5", "CF6": "cf6", "CF7": "cf7", "CF8": "cf8", "CF9": "cf9",
    "Potential DM's\nInternal": "damageMechanismInternal", "Potential DM's\nExternal": "damageMechanismExternal",
    "EL \nInternal": "elInternal", "EL \nExternal": "elExternal", "Min EL (Months)": "minELMonths",
    "Risk \n(1AP)": "risk1AP", "Risk \n(2AP)": "risk2AP", "Risk \n(3AP)": "risk3AP", "RLI \n(Months)": "rliMonths",
    "Criticality 1AP": "criticality1AP", "RLI Due Date\nmm/dd/yyyy": "rliDueDate",
    "Inspection Due Date\nmm/dd/yyyy": "inspectionDue", "Inspection Scope\nyyyy": "inspectionScope",
    "Note": "note", "Remarks": "remarks",
}

def clean(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return None
    return value

wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
headers = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
header_map = {str(v).strip(): i for i, v in enumerate(headers) if v not in (None, "")}

missing = [h for h in FIELDS if h not in header_map]
if missing:
    raise RuntimeError("Missing expected columns: " + ", ".join(missing))

records = []
for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
    first = row[header_map["No"]]
    if not isinstance(first, (int, float)):
        continue

    record = {}
    for header, key in FIELDS.items():
        record[key] = clean(row[header_map[header]])
    records.append(record)

if not records:
    raise RuntimeError("No RBI asset records found in Input RBI.xlsx")

OUTPUT.write_text(
    "window.RBI_DATA = " + json.dumps(records, ensure_ascii=False, separators=(",", ":")) + ";\n",
    encoding="utf-8",
)

print(f"Generated {OUTPUT} with {len(records)} RBI assets")
